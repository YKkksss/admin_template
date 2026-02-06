"""
Excel 导入/导出通用工具。

说明：
- 当前统一使用 xlsx（Office Open XML）格式。
- 该模块只负责“读写 Excel”，不耦合具体业务（用户/字典等）。
"""

from __future__ import annotations

from collections.abc import Callable
from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


@dataclass(frozen=True)
class ExcelColumn:
    """Excel 列定义。"""

    title: str
    key: str
    required: bool = False
    width: int | None = None
    formatter: Callable[[Any], Any] | None = None
    parser: Callable[[Any], Any] | None = None


@dataclass(frozen=True)
class ExcelImportError:
    """导入错误信息（用于返回给前端/日志记录）。"""

    row: int
    column: str | None
    message: str


def _to_cell_value(val: Any) -> Any:
    """将 Python 值转换为更适合写入 Excel 的值。"""

    if val is None:
        return None
    if isinstance(val, datetime):
        # 统一输出为本地时间字符串，避免时区与 Excel 日期格式差异带来的歧义
        return val.astimezone().strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(val, date):
        return val.isoformat()
    return val


def _auto_width(columns: list[ExcelColumn], rows: list[dict[str, Any]]) -> list[int]:
    """根据表头与少量数据估算列宽（字符数），避免生成的表格过窄。"""

    widths: list[int] = []
    sample_rows = rows[:200]  # 仅采样前 200 行，避免大数据集耗时
    for col in columns:
        if col.width:
            widths.append(int(col.width))
            continue

        max_len = len(str(col.title))
        for r in sample_rows:
            v = r.get(col.key)
            v = _to_cell_value(v)
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))

        # +2 留出边距，限制最大宽度避免异常值撑爆
        widths.append(min(max_len + 2, 60))

    return widths


def _write_sheet(
    ws,
    *,
    columns: list[ExcelColumn],
    rows: list[dict[str, Any]],
) -> None:
    """写入表头/数据/列宽等基础样式。"""

    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")

    # 表头
    for idx, col in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=idx, value=col.title)
        cell.font = header_font
        cell.alignment = header_alignment

    ws.freeze_panes = "A2"

    # 数据
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, col in enumerate(columns, start=1):
            v = row.get(col.key)
            if col.formatter:
                try:
                    v = col.formatter(v)
                except Exception:
                    # 格式化失败不应导致整个导出失败，兜底使用原值
                    pass
            ws.cell(row=row_idx, column=col_idx, value=_to_cell_value(v))

    # 列宽
    widths = _auto_width(columns, rows)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = float(w)

    # 自动筛选范围
    if columns:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{max(1, len(rows) + 1)}"


def export_xlsx_bytes(
    *,
    sheet_name: str,
    columns: list[ExcelColumn],
    rows: list[dict[str, Any]],
) -> bytes:
    """
    将数据导出为 xlsx 文件内容（bytes）。

    - rows：每行一个 dict，key 与 ExcelColumn.key 对应
    """

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31] if sheet_name else "Sheet1"
    _write_sheet(ws, columns=columns, rows=rows)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def append_sheet(
    wb: Workbook,
    *,
    sheet_name: str,
    columns: list[ExcelColumn],
    rows: list[dict[str, Any]],
) -> None:
    """向 Workbook 追加一个 sheet（用于导入模板的参考数据）。"""

    ws = wb.create_sheet(title=sheet_name[:31] if sheet_name else "Sheet")
    _write_sheet(ws, columns=columns, rows=rows)


def parse_xlsx_rows(
    *,
    content: bytes,
    columns: list[ExcelColumn],
    sheet_name: str | None = None,
) -> tuple[list[tuple[int, dict[str, Any]]], list[ExcelImportError]]:
    """
    解析 xlsx，返回（行号, 行数据）列表 + 错误列表。

    约定：
    - 第一行必须为表头
    - 允许 Excel 额外列存在（不参与解析）
    """

    wb = load_workbook(BytesIO(content), data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    errors: list[ExcelImportError] = []

    header_row = [c.value for c in ws[1]]
    header_titles = [str(v).strip() if v is not None else "" for v in header_row]
    title_to_index: dict[str, int] = {}
    for idx, title in enumerate(header_titles):
        if not title:
            continue
        title_to_index[title] = idx

    # 校验必须字段是否存在
    missing = [c.title for c in columns if c.required and c.title not in title_to_index]
    if missing:
        raise ValueError(f"缺少必填列：{', '.join(missing)}")

    parsed: list[tuple[int, dict[str, Any]]] = []
    for row_idx in range(2, ws.max_row + 1):
        values = [c.value for c in ws[row_idx]]
        # 空行直接跳过
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in values):
            continue

        row_data: dict[str, Any] = {}
        for col in columns:
            col_index = title_to_index.get(col.title)
            if col_index is None:
                # 非必填列缺失时，按 None 处理
                row_data[col.key] = None
                continue

            cell_val = values[col_index] if col_index < len(values) else None
            if isinstance(cell_val, str):
                cell_val = cell_val.strip()

            if col.required and (cell_val is None or cell_val == ""):
                errors.append(ExcelImportError(row=row_idx, column=col.title, message="不能为空"))
                continue

            if col.parser and cell_val not in (None, ""):
                try:
                    cell_val = col.parser(cell_val)
                except Exception as exc:  # noqa: BLE001 - 需要兜底记录错误
                    errors.append(
                        ExcelImportError(
                            row=row_idx,
                            column=col.title,
                            message=f"解析失败：{exc}",
                        ),
                    )
                    continue

            row_data[col.key] = cell_val

        parsed.append((row_idx, row_data))

    return parsed, errors
